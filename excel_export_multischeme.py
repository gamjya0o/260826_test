"""
Excel export for the CDR codon-optimization tool.

Produces a workbook where each mutant gets TWO rows under the FR/CDR-split
column layout (No. / Clone name / FR1_Vk / CDR1_Vk / ... / FR4_VH):
  - an amino-acid row (mutant FASTA, split into FR/CDR), with every residue
    that differs from WT shown in bold red
  - directly below it, the corresponding nucleotide row (codons for that
    same FR/CDR span), with every codon that was actually swapped
    (i.e. every CDR-region change) shown in bold red

Light-chain segments go under the *_Vk columns, heavy-chain segments go
under the *_VH columns (matches the lab's existing spreadsheet convention).
"""

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from cdr_codon_tool_multischeme import (
    FRAME_ORDER, region_of, which_cdr, build_kabat_to_seqidx, get_all_domains,
    DEFAULT_SCHEME,
)

NORMAL_FONT = InlineFont(rFont="Consolas", sz=10)
HIGHLIGHT_FONT = InlineFont(rFont="Consolas", sz=10, b=True, color="FF0000")

HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
LABEL_FONT = Font(name="Arial", size=10)
SEQ_FONT = Font(name="Consolas", size=10)
NT_ROW_FONT = Font(name="Arial", size=9, italic=True, color="666666")

COLUMNS = [
    "No.", "Clone name",
    "FR1_Vk", "CDR1_Vk", "FR2_Vk", "CDR2_Vk", "FR3_Vk", "CDR3_Vk", "FR4_Vk",
    "FR1_VH", "CDR1_VH", "FR2_VH", "CDR2_VH", "FR3_VH", "CDR3_VH", "FR4_VH",
    "Vk_full_length_AA", "Vk_full_length_NT",
    "VH_full_length_AA", "VH_full_length_NT",
]
VK_LABELS = ["FR1_Vk", "CDR1_Vk", "FR2_Vk", "CDR2_Vk", "FR3_Vk", "CDR3_Vk", "FR4_Vk"]
VH_LABELS = ["FR1_VH", "CDR1_VH", "FR2_VH", "CDR2_VH", "FR3_VH", "CDR3_VH", "FR4_VH"]


def _domain_segments_with_idx(seq, domain, scheme=DEFAULT_SCHEME, next_start=None):
    """Same as frame_one_domain, but keeps (char, absolute_seq_index) pairs
    per FR/CDR label instead of collapsing straight to a string.
    If `next_start` is given, everything from just after this domain's
    official end up to (not including) next_start is appended onto FR4 —
    i.e. a linker before the next domain, or a C-terminal tag/constant
    region after the last domain, gets folded into FR4 instead of being
    dropped."""
    numbering, chain_type, start = domain["numbering"], domain["chain_type"], domain["start"]
    pos2idx = build_kabat_to_seqidx(numbering, start)
    ordered = sorted(pos2idx.items(), key=lambda kv: (kv[0][0], kv[0][1]))
    segments = {name: [] for name in FRAME_ORDER}
    
    max_idx_used = -1
    for (num, ins), idx in ordered:
        seg = region_of(num, chain_type, scheme)
        segments[seg].append((seq[idx], idx))
        max_idx_used = max(max_idx_used, idx)
        
    # ANARCI 넘버링이 끝난 지점 이후의 꼬리 서열을 전부 FR4에 이어 붙임
    tail_start = max_idx_used + 1 if max_idx_used >= 0 else domain["start"]
    if next_start is not None and next_start > tail_start:
        for idx in range(tail_start, next_start):
            if idx < len(seq):
                segments["FR4"].append((seq[idx], idx))
                
    return segments, chain_type


def _rich_text(chars_with_flags):
    """chars_with_flags: list of (text_chunk, is_highlighted). Merges
    consecutive same-flag chunks and returns a CellRichText (or plain str
    if nothing is highlighted, for a clean-looking normal cell)."""
    if not chars_with_flags:
        return ""
    if not any(flag for _, flag in chars_with_flags):
        return "".join(t for t, _ in chars_with_flags)

    blocks = []
    cur_text, cur_flag = "", chars_with_flags[0][1]
    for t, flag in chars_with_flags:
        if flag == cur_flag:
            cur_text += t
        else:
            blocks.append((cur_text, cur_flag))
            cur_text, cur_flag = t, flag
    blocks.append((cur_text, cur_flag))

    parts = []
    for text, flag in blocks:
        if flag:
            parts.append(TextBlock(HIGHLIGHT_FONT, text))
        else:
            parts.append(TextBlock(NORMAL_FONT, text))
    return CellRichText(*parts)


def _aa_row_cells(mut_domains, mutant_aa_seq, mutated_mut_idx_set, scheme=DEFAULT_SCHEME):
    """Returns {column_label: cell_value} for the amino-acid row."""
    out = {}
    for i, dom in enumerate(mut_domains):
        next_start = mut_domains[i + 1]["start"] if i + 1 < len(mut_domains) else len(mutant_aa_seq)
        segments, chain_type = _domain_segments_with_idx(mutant_aa_seq, dom, scheme, next_start)
        labels = VK_LABELS if chain_type in ("K", "L") else VH_LABELS
        for frame_name, col_label in zip(FRAME_ORDER, labels):
            chars = [(c, idx in mutated_mut_idx_set) for c, idx in segments[frame_name]]
            out[col_label] = _rich_text(chars)
    return out


def _nt_row_cells(wt_domains, result):
    """Returns {column_label: cell_value} for the nucleotide row.
    Indel-aware: uses codon_list/deleted_wt_idx/insertions_after so that
    deleted CDR codons disappear and inserted CDR codons are spliced in,
    instead of naively slicing a fixed-length final sequence."""
    scheme = result.get("scheme", DEFAULT_SCHEME)
    codon_list = result["codon_list"]
    deleted = result["deleted_wt_idx"]
    insertions_after = result["insertions_after"]
    applied_wt_idx_set = {m["wt_seq_index"] for m in result["cdr_mutations_applied"] if m["type"] == "substitution"}

    out = {}
    for i, dom in enumerate(wt_domains):
        next_start = wt_domains[i + 1]["start"] if i + 1 < len(wt_domains) else len(codon_list)
        placeholder_len = max(dom["end"], next_start - 1, 0) + 1
        dummy_seq = "?" * placeholder_len if dom["end"] >= 0 else ""
        segments, chain_type = _domain_segments_with_idx(dummy_seq, dom, scheme, next_start)
        labels = VK_LABELS if chain_type in ("K", "L") else VH_LABELS
        for frame_name, col_label in zip(FRAME_ORDER, labels):
            chunks = []
            for _, idx in segments[frame_name]:
                if idx not in deleted:
                    chunks.append((codon_list[idx], idx in applied_wt_idx_set))
                for ins_codon in insertions_after.get(idx, []):
                    chunks.append((ins_codon, True))
            out[col_label] = _rich_text(chunks)
    return out


def _domain_full_region_idx(domain, scheme=DEFAULT_SCHEME, next_start=None, placeholder_len=None):
    """Return [(char_placeholder, abs_idx), ...] for the WHOLE V-region of one
    domain (FR1..FR4 concatenated, in sequence order). If `next_start` is
    given, everything up to (not including) next_start — a linker or
    C-terminal tag/constant region — is folded into FR4 too."""
    plen = placeholder_len if placeholder_len is not None else domain["end"]
    if next_start is not None:
        plen = max(plen, next_start - 1)
    dummy_seq = "?" * (plen + 1)
    segments, chain_type = _domain_segments_with_idx(dummy_seq, domain, scheme, next_start)
    combined = []
    for frame in FRAME_ORDER:
        combined.extend(segments[frame])  # (char, idx) but char is placeholder here
    return [idx for _, idx in combined], chain_type


def _vregion_aa_rich(domain, seq, highlight_idx_set, scheme=DEFAULT_SCHEME, next_start=None):
    idxs, chain_type = _domain_full_region_idx(domain, scheme, next_start=next_start)
    chars = [(seq[idx], idx in highlight_idx_set) for idx in idxs]
    return _rich_text(chars), chain_type


def _vregion_nt_rich(domain, result, next_start=None):
    scheme = result.get("scheme", DEFAULT_SCHEME)
    idxs, chain_type = _domain_full_region_idx(domain, scheme, next_start=next_start)
    codon_list = result["codon_list"]
    deleted = result["deleted_wt_idx"]
    insertions_after = result["insertions_after"]
    applied_wt_idx_set = {m["wt_seq_index"] for m in result["cdr_mutations_applied"] if m["type"] == "substitution"}

    chunks = []
    for idx in idxs:
        if idx not in deleted:
            chunks.append((codon_list[idx], idx in applied_wt_idx_set))
        for ins_codon in insertions_after.get(idx, []):
            chunks.append((ins_codon, True))
    return _rich_text(chunks), chain_type


def _wt_dummy_result(wt, scheme=DEFAULT_SCHEME):
    """Build a minimal 'result'-like dict for the WT reference row (no mutations)."""
    return {
        "scheme": scheme,
        "codon_list": [wt["nt"][i:i + 3] for i in range(0, len(wt["nt"]), 3)],
        "deleted_wt_idx": set(),
        "insertions_after": {},
        "cdr_mutations_applied": [],
    }


def build_workbook(wt_pairs, mutant_results, output_path, scheme=DEFAULT_SCHEME):
    """
    wt_pairs: list of WT dicts (id, aa, nt, ...) as produced by extract_wt_pairs
    mutant_results: list of (mutant_id, mutant_aa_seq, analyze()-result-dict, matched_wt)
    scheme: numbering scheme used (affects only CDR/FR boundaries) — should match
    the scheme used when wt_pairs/mutant_results were generated.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"CDR_codon_result_{scheme}"

    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    row_idx = 2

    # WT reference row(s)
    for wt in wt_pairs:
        wt_domains = get_all_domains(wt["aa"], wt["id"], scheme=scheme)
        ws.cell(row=row_idx, column=1, value="-").font = LABEL_FONT
        ws.cell(row=row_idx, column=2, value=f"WT: {wt['id']}").font = Font(name="Arial", size=10, bold=True)
        cells = _aa_row_cells(wt_domains, wt["aa"], set(), scheme)
        for col_label, value in cells.items():
            col_idx = COLUMNS.index(col_label) + 1
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font = SEQ_FONT
        for i, dom in enumerate(wt_domains):
            next_start = wt_domains[i + 1]["start"] if i + 1 < len(wt_domains) else len(wt["aa"])
            aa_rich, chain_type = _vregion_aa_rich(dom, wt["aa"], set(), scheme, next_start=next_start)
            nt_rich, _ = _vregion_nt_rich(dom, _wt_dummy_result(wt, scheme), next_start=next_start)
            aa_col = "Vk_full_length_AA" if chain_type in ("K", "L") else "VH_full_length_AA"
            nt_col = "Vk_full_length_NT" if chain_type in ("K", "L") else "VH_full_length_NT"
            ws.cell(row=row_idx, column=COLUMNS.index(aa_col) + 1, value=aa_rich).font = SEQ_FONT
            ws.cell(row=row_idx, column=COLUMNS.index(nt_col) + 1, value=nt_rich).font = SEQ_FONT
        row_idx += 1

    ws.cell(row=row_idx, column=1, value="").font = LABEL_FONT
    row_idx += 1  # blank spacer row

    for mutant_id, mutant_aa_seq, result, wt in mutant_results:
        no_val = mutant_id

        mutated_mut_idx_set = {m["mut_seq_index"] for m in result["mutations"] if m["mut_seq_index"] is not None}

        # --- amino acid row ---
        ws.cell(row=row_idx, column=1, value=no_val).font = LABEL_FONT
        ws.cell(row=row_idx, column=2, value=mutant_id).font = Font(name="Arial", size=10, bold=True)
        aa_cells = _aa_row_cells(result["mut_domains"], mutant_aa_seq, mutated_mut_idx_set, result.get("scheme", scheme))
        for col_label, value in aa_cells.items():
            col_idx = COLUMNS.index(col_label) + 1
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font = SEQ_FONT
        for i, dom in enumerate(result["mut_domains"]):
            next_start = result["mut_domains"][i + 1]["start"] if i + 1 < len(result["mut_domains"]) else len(mutant_aa_seq)
            aa_rich, chain_type = _vregion_aa_rich(dom, mutant_aa_seq, mutated_mut_idx_set, result.get("scheme", scheme), next_start=next_start)
            aa_col = "Vk_full_length_AA" if chain_type in ("K", "L") else "VH_full_length_AA"
            ws.cell(row=row_idx, column=COLUMNS.index(aa_col) + 1, value=aa_rich).font = SEQ_FONT
        aa_row = row_idx
        row_idx += 1

        # --- nucleotide row (directly below) ---
        ws.cell(row=row_idx, column=2, value=f"{mutant_id} (nucleotide)").font = NT_ROW_FONT
        nt_cells = _nt_row_cells(result["wt_domains"], result)
        for col_label, value in nt_cells.items():
            col_idx = COLUMNS.index(col_label) + 1
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font = SEQ_FONT
            c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for i, dom in enumerate(result["wt_domains"]):
            next_start = result["wt_domains"][i + 1]["start"] if i + 1 < len(result["wt_domains"]) else len(result["codon_list"])
            nt_rich, chain_type = _vregion_nt_rich(dom, result, next_start=next_start)
            nt_col = "Vk_full_length_NT" if chain_type in ("K", "L") else "VH_full_length_NT"
            c = ws.cell(row=row_idx, column=COLUMNS.index(nt_col) + 1, value=nt_rich)
            c.font = SEQ_FONT
            c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        row_idx += 1
        row_idx += 1  # blank spacer row between mutants

    # column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    for col_label in VK_LABELS + VH_LABELS:
        col_idx = COLUMNS.index(col_label) + 1
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    for col_label in ["Vk_full_length_AA", "VH_full_length_AA"]:
        ws.column_dimensions[get_column_letter(COLUMNS.index(col_label) + 1)].width = 45
    for col_label in ["Vk_full_length_NT", "VH_full_length_NT"]:
        ws.column_dimensions[get_column_letter(COLUMNS.index(col_label) + 1)].width = 65

    wb.save(output_path)
    return output_path
